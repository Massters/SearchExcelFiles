import sys
import os
import re
import openpyxl
from queue import Queue
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QInputDialog, QVBoxLayout, QWidget, QLabel, QTableWidget, QTableWidgetItem, QHBoxLayout, QAction, QCheckBox, QMessageBox, QDialog, QCheckBox, QPushButton, QLineEdit, QProgressBar
from PyQt5.QtCore import Qt, QObject, QRunnable, QThreadPool, pyqtSignal

# 定义了自定义的信号用于在Excel搜索任务中发射信号
class ExcelSearchTaskSignals(QObject):
    foundKeyword = pyqtSignal(str, str, int, int, str)
    foundAdditionalContent = pyqtSignal(str)
    finished = pyqtSignal(str)
    error = pyqtSignal(str)

# 表示一个Excel搜索任务.在后台线程中执行实际的Excel搜索操作,并通过信号将搜索结果传递至主窗口
class ExcelSearchTask(QRunnable):
    def __init__(self, file_path, keyword, use_regex):
        super().__init__()
        self.file_path = file_path
        self.keyword = keyword
        self.use_regex = use_regex
        self.signals = ExcelSearchTaskSignals()

    def run(self):
        global error_occurred
        try:
            if error_occurred:
                return
            else:
                workbook = openpyxl.load_workbook(self.file_path, read_only=True)

                for sheet_name in workbook.sheetnames:
                    sheet = workbook[sheet_name]
                    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                        found_keyword = False
                        additional_content_list = []
                        for col_idx, cell in enumerate(row, start=1):
                            cell_value = str(cell)
                            if self.use_regex:
                                pattern = re.compile(self.keyword, re.IGNORECASE)
                                if pattern.search(cell_value):
                                    found_keyword = True
                                    self.signals.foundKeyword.emit(self.file_path, sheet_name, col_idx, row_idx, cell_value)
                            else:
                                if self.keyword in cell_value:
                                    found_keyword = True
                                    self.signals.foundKeyword.emit(self.file_path, sheet_name, col_idx, row_idx, cell_value)

                            if found_keyword:
                                found_keyword = False
                                for cell in row:
                                    cell_value = str(cell)
                                    additional_content_list.append(cell_value)
                                additional_content = " ".join(additional_content_list).replace("None", " ")
                                additional_content_list = []
                                self.signals.foundAdditionalContent.emit(additional_content.strip())

                self.signals.finished.emit(self.file_path)

                workbook.close()
        except Exception as e:
            error_occurred = True
            self.signals.error.emit(str(e))

# 表示应用程序的主窗口,它包含了用户界面的布局, 控件和与搜索任务相关的处理方法
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.initUI()

    # 定义了程序的用户界面布局,
    def initUI(self):
        mainWidget = QWidget()
        mainLayout = QVBoxLayout()

        # titleLabel = QLabel("Excel Search")
        # titleLabel.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 10px;")

        # 定义表格用于显示搜索结果
        self.tableWidget = QTableWidget()
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels(["File", "Sheet", "Cell", "Content", "Additional Content"])
        self.tableWidget.setColumnWidth(0, 1000) #设置 File 标签的宽度
        self.tableWidget.setStyleSheet("border: 1px solid black;")
        self.tableWidget.horizontalHeader().setStretchLastSection(True)

        # mainLayout.addWidget(titleLabel)
        mainLayout.addWidget(self.tableWidget)

        self.progressLabel = QLabel()
        mainLayout.addWidget(self.progressLabel)

        self.progressBar = QProgressBar()
        mainLayout.addWidget(self.progressBar)

        mainWidget.setLayout(mainLayout)
        self.setCentralWidget(mainWidget)

        # 定义菜单栏用于打开文件夹
        openFolder = QAction('Open Folder', self)
        openFolder.setShortcut('Ctrl+O')
        openFolder.triggered.connect(self.openFolderDialog)

        self.menuBar().addMenu('File').addAction(openFolder)

        self.threadpool = QThreadPool()
        self.threadpool.setMaxThreadCount(4)
        self.task_queue = Queue()

        self.setGeometry(300, 300, 2500, 1000)
        self.setWindowTitle('Excel Search')
        self.show()

    # 打开文件夹对话框
    def openFolderDialog(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.DirectoryOnly)
        dialog.setWindowTitle('Open Folder')

        if dialog.exec_() == QFileDialog.Accepted:
            folder_path = dialog.selectedFiles()[0] #获取路径
            self.searchExcelFiles(folder_path)

    ################################################################################################################################################################
    # 遍历选定的文件夹中所有的Excel文件,并为每个文件创建一个ExcelSearchTask实例, 然后将其添加到线程池中,通过连接信号和槽函数, 搜索任务的结果将传递给handleKeywordFound和
    # handleAdditionalContentFound方法,这些方法负责将结果显示在表格中
    ################################################################################################################################################################
    def searchExcelFiles(self, folder_path):
        dialog = QDialog()
        dialog.setWindowTitle('Enter Keyword')

        layout = QVBoxLayout()
        dialog.setLayout(layout)

        keyword_input = QLineEdit()
        layout.addWidget(keyword_input)

        checkbox_re = QCheckBox("使用正则表达式")
        layout.addWidget(checkbox_re)

        ok_button = QPushButton("OK")
        ok_button.clicked.connect(lambda: self.onSearch(dialog, keyword_input.text(), folder_path, checkbox_re.isChecked()))
        layout.addWidget(ok_button)

        dialog.exec_()

    def onSearch(self, dialog, keyword, folder_path, use_regex):
        if keyword == "":
            QMessageBox.warning(self, "ERROR", "Please Enter Keyword!")
        else:
            dialog.close()
            # 清除旧结果
            global insert_row
            insert_row = 0
            self.tableWidget.setRowCount(0)
            self.threadpool.clear()

            global error_occurred 
            error_occurred = False
            total_threads = 0
            self.completed_threads = 0

            self.progressBar.setValue(0)

            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.endswith(".xlsx") or file.endswith(".xlsm"):
                        total_threads += 1
                        file_path = root + "/" + file
                        task = ExcelSearchTask(file_path, keyword, use_regex)
                        task.signals.foundKeyword.connect(self.handleKeywordFound)
                        task.signals.foundAdditionalContent.connect(self.handleAdditionalContentFound)
                        task.signals.finished.connect(self.handleTaskFinished)
                        task.signals.error.connect(self.handleTaskError)
                        self.task_queue.put(task)

            self.total_threads = total_threads
            self.progressLabel.setText(f"Searching files: 0 / {total_threads}")

            while not self.task_queue.empty():
                task = self.task_queue.get()
                self.threadpool.start(task)

    # 用于处理在Excel文件中找到的关键字事件, 它将搜索结果添加到表格中的相应单元格
    def handleKeywordFound(self, file_path, sheet_name, col_idx, row_idx, cell_value):
        global insert_row
        self.tableWidget.insertRow(insert_row)

        self.tableWidget.setItem(insert_row, 0, QTableWidgetItem(file_path))
        self.tableWidget.setItem(insert_row, 1, QTableWidgetItem(sheet_name))
        self.tableWidget.setItem(insert_row, 2, QTableWidgetItem(f"{get_column_letter(col_idx)}{row_idx}"))
        self.tableWidget.setItem(insert_row, 3, QTableWidgetItem(cell_value))

    # 用于处理在Excel文件中找到的附加内容事件, 它将附加内容添加到表格中的相应单元格
    def handleAdditionalContentFound(self, additional_content):
        # row_count = self.tableWidget.rowCount()
        global insert_row
        self.tableWidget.setItem(insert_row, 4, QTableWidgetItem(additional_content))
        insert_row += 1

    # 用于处理搜索任务完成事件
    def handleTaskFinished(self, file_path):
        self.completed_threads += 1

        self.progressLabel.setText(f"Searching files: {self.completed_threads} / {self.total_threads}\t  {file_path}")
        self.progressBar.setValue(int(self.completed_threads * 100 / self.total_threads))

        if self.completed_threads == self.total_threads:
            result_count = self.tableWidget.rowCount()
            QMessageBox.information(self, "Results", f"共找到了{self.tableWidget.rowCount()}个结果")

    # 用于处理搜索任务发生错误的事件, 它显示一个错误的消息框
    def handleTaskError(self, error_msg):
        # global error_occurred
        # error_occurred = True
        self.threadpool.clear()
        self.tableWidget.setRowCount(0)
        self.progressLabel.setText("Searching files: Error occurred!")
        self.progressBar.setValue(0)

        QMessageBox.critical(self, "Error", error_msg)

error_occurred = False
insert_row = 0

if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = MainWindow()
    sys.exit(app.exec_())