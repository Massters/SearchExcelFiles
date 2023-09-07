import sys
import os
import re
import openpyxl
from openpyxl.utils import get_column_letter
from PyQt5.QtWidgets import QApplication, QMainWindow, QFileDialog, QInputDialog, QVBoxLayout, QWidget, QLabel, QTableWidget, QTableWidgetItem, QHBoxLayout, QAction, QCheckBox, QMessageBox, QDialog,QCheckBox, QPushButton,QLineEdit

class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.regexCheckbox = None
        self.initUI()

    def initUI(self):
        mainWidget = QWidget()
        mainLayout = QVBoxLayout()

        titleLabel = QLabel("Excel Search")
        titleLabel.setStyleSheet("font-size: 20px; font-weight: bold; margin-bottom: 10px;")

        self.tableWidget = QTableWidget()
        self.tableWidget.setColumnCount(5)
        self.tableWidget.setHorizontalHeaderLabels(["File", "Sheet", "Cell", "Content", "Additional Content"])
        self.tableWidget.setStyleSheet("border: 1px solid black;")
        self.tableWidget.horizontalHeader().setStretchLastSection(True)

        mainLayout.addWidget(titleLabel)
        mainLayout.addWidget(self.tableWidget)

        mainWidget.setLayout(mainLayout)
        self.setCentralWidget(mainWidget)

        openFolder = QAction('Open Folder', self)
        openFolder.setShortcut('Ctrl+O')
        openFolder.triggered.connect(self.openFolderDialog)

        self.menuBar().addMenu('File').addAction(openFolder)

        self.setGeometry(300, 300, 800, 600)
        self.setWindowTitle('Excel Search')
        self.show()

    def openFolderDialog(self):
        dialog = QFileDialog()
        dialog.setFileMode(QFileDialog.DirectoryOnly)
        dialog.setWindowTitle('Open Folder')

        if dialog.exec_() == QFileDialog.Accepted:
            folder_path = dialog.selectedFiles()[0]
            self.searchExcelFiles(folder_path)

    def searchExcelFiles(self, folder_path):
        dialog = QDialog()
        dialog.setWindowTitle('Enter Keyword')

        layout = QVBoxLayout()
        dialog.setLayout(layout)

        keyword_input = QLineEdit()
        layout.addWidget(keyword_input)

        checkbox_re = QCheckBox("使用正则表达式")
        layout.addWidget(checkbox_re)

        ok_button = QPushButton("OK")
        ok_button.clicked.connect(lambda: self.onSearch(dialog, keyword_input.text(), folder_path,checkbox_re.isChecked()))
        layout.addWidget(ok_button)
        cancel_button = QPushButton('Cancel')
        cancel_button.clicked.connect(dialog.reject)
        layout.addWidget(cancel_button)

        dialog.exec_()

    def onSearch(self, dialog, keyword, folder_path, checkbox_re_checked):
        dialog.close()
        
        self.tableWidget.clearContents()
        self.tableWidget.setRowCount(0)

        if len(keyword) < 1:
            QMessageBox.warning(self,"Error", "请输入Keyword")
        else:
            for root, dirs, files in os.walk(folder_path):
                for file in files:
                    if file.endswith(".xlsx") or file.endswith(".xlsm"):
                        file_path = os.path.join(root, file)
                        if checkbox_re_checked:
                            try:
                                self.searchExcelFileWithRegex(file_path, keyword)
                            except Exception as e:
                                QMessageBox.warning(self,"Error", str(e))
                                return
                        else:
                            self.searchExcelFile(file_path, keyword)

    def searchExcelFile(self, file_path, keyword):
        workbook = openpyxl.load_workbook(file_path, read_only=True)

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                found_keyword = False
                additional_content_list = []
                for col_idx, cell in enumerate(row, start=1):
                    cell_value = str(cell)
                    if isinstance(cell_value, str) and keyword in cell_value:
                        found_keyword = True
                        file_item = QTableWidgetItem(file_path)
                        sheet_item = QTableWidgetItem(sheet_name)
                        cell_position = QTableWidgetItem(f"{get_column_letter(col_idx)}{row_idx}")
                        content_item = QTableWidgetItem(cell_value)
                        additional_content_item = QTableWidgetItem("")

                        row_count = self.tableWidget.rowCount()
                        self.tableWidget.insertRow(row_count)
                        self.tableWidget.setItem(row_count, 0, file_item)
                        self.tableWidget.setItem(row_count, 1, sheet_item)
                        self.tableWidget.setItem(row_count, 2, cell_position)
                        self.tableWidget.setItem(row_count, 3, content_item)
                        self.tableWidget.setItem(row_count, 4, additional_content_item)

                    if col_idx != 1:
                        additional_content_list.append(cell_value)

                if found_keyword:
                    additional_content = " ".join(additional_content_list)
                    additional_content_item.setText(additional_content.strip())

        workbook.close()
    
    def searchExcelFileWithRegex(self, file_path, keyword):
        self.error_shown = False

        try:
            pattern = re.compile(keyword, re.IGNORECASE)  # 创建正则表达式模式
            
            workbook = openpyxl.load_workbook(file_path, read_only=True)

            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                    found_keyword = False
                    additional_content_list = []
                    for col_idx, cell in enumerate(row, start=1):
                        cell_value = str(cell)
                        if pattern.search(cell_value):
                            found_keyword = True
                            file_item = QTableWidgetItem(file_path)
                            sheet_item = QTableWidgetItem(sheet_name)
                            cell_position = QTableWidgetItem(f"{get_column_letter(col_idx)}{row_idx}")
                            content_item = QTableWidgetItem(cell_value)
                            additional_content_item = QTableWidgetItem("")

                            row_count = self.tableWidget.rowCount()
                            self.tableWidget.insertRow(row_count)
                            self.tableWidget.setItem(row_count, 0, file_item)
                            self.tableWidget.setItem(row_count, 1, sheet_item)
                            self.tableWidget.setItem(row_count, 2, cell_position)
                            self.tableWidget.setItem(row_count, 3, content_item)
                            self.tableWidget.setItem(row_count, 4, additional_content_item)

                        if col_idx != 1:
                            additional_content_list.append(cell_value)

                    if found_keyword:
                        additional_content = " ".join(additional_content_list)
                        additional_content_item.setText(additional_content.strip())

            workbook.close()
        except re.error as e:
            raise

if __name__ == '__main__':
    app = QApplication(sys.argv)
    mainWindow = MainWindow()
    sys.exit(app.exec_())